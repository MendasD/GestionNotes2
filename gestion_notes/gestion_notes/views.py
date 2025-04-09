from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404, FileResponse
from urllib.parse import quote  
from django.contrib import messages
import pandas as pd
from connexion.models import Etudiant, Note, Matiere, Classe, Responsable, Message, FichiersJoints, Moyenne, Enseignants, EmploiDuTemps, Programmation_cours
from .forms import NoteUploadForm, MatiereForm, ResponsableForm, EtudiantForm, ClasseForm, EtudiantUpdateForm, OneMessageForm, EtudiantUpdatePasswordForm, EnseignantForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from .mes_decorateurs import is_login
from django.utils import timezone
from django.contrib.auth.hashers import check_password,make_password
import json
import os
import openpyxl 
from openpyxl import Workbook
import io
import xlwings as xw
from tempfile import NamedTemporaryFile
import numpy as np
from django.db.models import Avg, Count
from django.utils import timezone, timesince
from datetime import timedelta
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import plotly.graph_objs as go # pour les graphiques interactifs
import plotly.offline as opy




# Verifie que le paramètre est un numérique compris entre  0 et 1
def is_float_between_0_and_1(value):
    """ Vérifie que la valeur est un float compris entre 0 et 1 """
    try:
        val = float(value)
    except ValueError:
        return False
    return 0 <= val <= 1

# Verifie que le paramètre est un numérique compris entre  0 et 1
def is_float_between_0_and_20(value):
    """ Vérifie que la valeur est un float positif entre 0 et 20"""
    try:
        val = float(value)
    except ValueError:
        return False
    return 0 <= val <= 20

def supprimer_notes_deja_ajoutees(matiere, annee_scolaire, classe, compteur):
    """ En cas d_erreurs lors de l_ajout des notes, 
    Permet de supprimer les notes déjà ajoutées jusque là
    
    attrs : 
        compteur : compte le nombre de notes déjà ajoutées
        matiere : la matière concernée
        annee_scolaire : l'année scolaire concernée
        classe : la classe concernée
    """

    notes = Note.objects.filter(matiere = matiere, annee_scolaire = annee_scolaire, classe = classe).order_by('-id')
    notes_a_supprimer = notes[:compteur]
    for note in notes_a_supprimer:
        note.delete()


@csrf_exempt
def ajouter_notes(request):
    """
        Cette fonction permet d'ajouter les notes des etudiants
        à partir d'un fichier Excel que l'on va parcourir ligne par ligne
    """
    if request.method == 'POST':
        try:
            data = request.POST  # Utilisation de request.POST pour les données textuelles
            fichier = request.FILES.get('fichier_excel')  # Récupération du fichier excel

            classeName = data.get('classe',"")
            semestre = data.get('semestre', "")
            matiereId = data.get('matiere', "")
            annee_scolaire = data.get('annee_scolaire', "")

            classe = Classe.objects.get(name=classeName)
            matiere = Matiere.objects.get(id=matiereId)

            fichier_is_valid = True

            try:
                df = pd.read_excel(fichier)

                # On vérifie d'abord la conformité des données du fichier excel

                for index, row in df.iterrows():
                    matricule = row['matricule']
                    try:
                        etudiant = Etudiant.objects.get(matricule=matricule)
                    except Etudiant.DoesNotExist:
                        fichier_is_valid = False
                        messages.error(request, f"L'étudiant avec le matricule {matricule} n'existe pas.")
                        print("erreur 1")
                        return JsonResponse({'status': 'error', 'message': f"L'étudiant avec le matricule {matricule} n'existe pas."}, status=404)


                    # Création des instances de Note pour chaque type de note
                    if 'note1' in df.columns:
                        try:
                            note2 = Note.objects.get(etudiant=etudiant,matiere=matiere, type_note='note2',annee_scolaire=annee_scolaire)
                            print("erreur test")
                            poids2 = note2.poids

                            if 'poids1' in df.columns:
                                poids1 = row['poids1']
                                if not is_float_between_0_and_1(poids1):

                                    # On supprime les notes déjà ajoutées jusque là
                                    #supprimer_notes_deja_ajoutees(matiere, annee_scolaire, classe, compteur)
                                    fichier_is_valid = False


                                    messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                else:
                                    if poids1 + poids2 != float(1):

                                        # On supprime les notes déjà ajoutées jusque là
                                        #supprimer_notes_deja_ajoutees(matiere, annee_scolaire, classe, compteur)
                                        fichier_is_valid = False

                                        messages.error(request, "La somme des poids des notes n'est pas égale à 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "La somme des poids des notes n'est pas égale à 1."}, status=400)
                                    else:
                                        if not is_float_between_0_and_20(row['note1']):
                                            fichier_is_valid = False
                                            messages.error(request, "La note doit être un réel entre 0 et 20")
                                            return JsonResponse({'status': 'error', 'message': "La note doit être un réel entre 0 et 20"}, status=400)
                                       
                            else:
                                # On supprime les notes déjà ajoutées jusque là
                                #supprimer_notes_deja_ajoutees(matiere, annee_scolaire, classe, compteur)
                                fichier_is_valid = False

                                messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                print("erreur 2")
                                return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                        except Note.DoesNotExist:
                            if 'poids1' in df.columns:
                                poids1 = row['poids1']
                                if not is_float_between_0_and_1(poids1):

                                    # On supprime les notes déjà ajoutées jusque là
                                    #supprimer_notes_deja_ajoutees(matiere, annee_scolaire, classe, compteur)
                                    fichier_is_valid = False

                                    messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                else:
                                    if 'poids2' in df.columns:
                                        if is_float_between_0_and_1(row['poids2']):
                                            if poids1 + row['poids2'] != float(1):
                                                fichier_is_valid = False
                                                messages.error(request, "La somme des deux poids ne donne pas 1")
                                                return JsonResponse({'status': 'error', 'message': "La somme des deux poids ne donne pas 1"}, status=400)
                                    
                                    if not is_float_between_0_and_20(row['note1']):
                                        fichier_is_valid = False
                                        messages.error(request, "La note doit être un réel entre 0 et 20")
                                        return JsonResponse({'status': 'error', 'message': "La note doit être un réel entre 0 et 20"}, status=400)
                            else:
                                fichier_is_valid = False

                                messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                print("erreur 2")
                                return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                    if 'note2' in df.columns:
                        try:
                            note1 = Note.objects.get(etudiant=etudiant,matiere=matiere, type_note='note1',annee_scolaire=annee_scolaire)
                            poids1 = note1.poids

                            if 'poids2' in df.columns:
                                poids2 = row['poids2']
                                if not is_float_between_0_and_1(poids2):
                                    fichier_is_valid = False

                                    messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                else:
                                    if poids1 + poids2 != float(1):
                                        fichier_is_valid = False

                                        messages.error(request, "La somme des poids des notes n'est pas égale à 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "La somme des poids des notes n'est pas égale à 1."}, status=400)
                                    else:
                                       if not is_float_between_0_and_20(row['note2']):
                                            fichier_is_valid = False
                                            messages.error(request, "La note doit être un réel entre 0 et 20")
                                            return JsonResponse({'status': 'error', 'message': "La note doit être un réel entre 0 et 20"}, status=400)
                            else:
                                fichier_is_valid = False

                                messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                print("erreur 2")
                                return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                        except Note.DoesNotExist:
                            if 'poids2' in df.columns:
                                poids2 = row['poids2']
                                if not is_float_between_0_and_1(poids2):
                                    fichier_is_valid = False
                                        
                                    messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                else:
                                   if not is_float_between_0_and_20(row['note2']):
                                        fichier_is_valid = False
                                        messages.error(request, "La note doit être un réel entre 0 et 20")
                                        return JsonResponse({'status': 'error', 'message': "La note doit être un réel entre 0 et 20"}, status=400)
                            else:
                                fichier_is_valid = False

                                messages.error(request, "Le poids de la note 2 n'est pas spécifié.")
                                print("erreur 2")
                                return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)
                    
                if fichier_is_valid:

                    for index, row in df.iterrows():
                        matricule = row['matricule']
                        try:
                            etudiant = Etudiant.objects.get(matricule=matricule)
                        except Etudiant.DoesNotExist:
                            messages.error(request, f"L'étudiant avec le matricule {matricule} n'existe pas.")
                            print("erreur 1")
                            return JsonResponse({'status': 'error', 'message': f"L'étudiant avec le matricule {matricule} n'existe pas."}, status=404)
            
                        # Création des instances de Note pour chaque type de note
                        if 'note1' in df.columns:
                            try:
                                note2 = Note.objects.get(etudiant=etudiant,matiere=matiere, type_note='note2',annee_scolaire=annee_scolaire)
                                print("erreur test")
                                poids2 = note2.poids

                                if 'poids1' in df.columns:
                                    poids1 = row['poids1']
                                    if not is_float_between_0_and_1(poids1):

                                        messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                    else:
                                        if poids1 + poids2 != float(1):

                                            messages.error(request, "La somme des poids des notes n'est pas égale à 1.")
                                            print("erreur 2")
                                            return JsonResponse({'status': 'error', 'message': "La somme des poids des notes n'est pas égale à 1."}, status=400)
                                        else:
                                            note1 = Note.objects.create(
                                                etudiant=etudiant,
                                                matiere=matiere,
                                                classe=classe,
                                                note=row['note1'],
                                                type_note='note1',
                                                poids = row['poids1'],
                                                semestre=semestre,
                                                annee_scolaire=annee_scolaire
                                            )

                                            moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=annee_scolaire).first()
                                            moyenne.moyenne = (note1.note * note1.poids + note2.note * note2.poids) 
                                            moyenne.save()
                                else:
                                   
                                    messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                            except Note.DoesNotExist:
                                if 'poids1' in df.columns:
                                    poids1 = row['poids1']
                                    if not is_float_between_0_and_1(poids1):

                                        messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                    else:
                                        note1 = Note.objects.create(
                                            etudiant=etudiant,
                                            matiere=matiere,
                                            classe=classe,
                                            note=row['note1'],
                                            type_note='note1',
                                            poids = row['poids1'],
                                            semestre=semestre,
                                            annee_scolaire=annee_scolaire
                                        )

                                        moyenne = Moyenne.objects.create(etudiant=etudiant, matiere=matiere, annee_scolaire=annee_scolaire, classe=classe,moyenne=row['note1']*poids1)
                                        
                                else:

                                    messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                            

                        if 'note2' in df.columns:
                            try:
                                note1 = Note.objects.get(etudiant=etudiant,matiere=matiere, type_note='note1',annee_scolaire=annee_scolaire)
                                poids1 = note1.poids

                                if 'poids2' in df.columns:
                                    poids2 = row['poids2']
                                    if not is_float_between_0_and_1(poids2):

                                        messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                    else:
                                        if poids1 + poids2 != float(1):

                                            messages.error(request, "La somme des poids des notes n'est pas égale à 1.")
                                            print("erreur 2")
                                            return JsonResponse({'status': 'error', 'message': "La somme des poids des notes n'est pas égale à 1."}, status=400)
                                        else:
                                            note2 = Note.objects.create(
                                                etudiant=etudiant,
                                                matiere=matiere,
                                                classe=classe,
                                                note=row['note2'],
                                                type_note='note2',
                                                poids = row['poids2'],
                                                semestre=semestre,
                                                annee_scolaire=annee_scolaire
                                            )

                                            moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=annee_scolaire).first()
                                            moyenne.moyenne = (note1.note * note1.poids + note2.note * note2.poids) 
                                            moyenne.save()
                                else:
                                    messages.error(request, "Le poids de la note 1 n'est pas spécifié.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                            except Note.DoesNotExist:
                                if 'poids2' in df.columns:
                                    poids2 = row['poids2']
                                    if not is_float_between_0_and_1(poids2):
                                        
                                        messages.error(request, "Le poids doit être un réel compris entre 0 et 1.")
                                        print("erreur 2")
                                        return JsonResponse({'status': 'error', 'message': "Le poids doit être compris entre 0 et 1.."}, status=400)
                                    else:
                                        note2 = Note.objects.create(
                                            etudiant=etudiant,
                                            matiere=matiere,
                                            classe=classe,
                                            note=row['note2'],
                                            type_note='note2',
                                            poids = row['poids2'],
                                            semestre=semestre,
                                            annee_scolaire=annee_scolaire
                                        )
                                        moyenne = Moyenne.objects.create(etudiant=etudiant, matiere=matiere, annee_scolaire=annee_scolaire, classe=classe,moyenne=row['note2']*poids2)
                                        
                                else:
                                    messages.error(request, "Le poids de la note 2 n'est pas spécifié.")
                                    print("erreur 2")
                                    return JsonResponse({'status': 'error', 'message': "Le poids de la note 1 n'est pas spécifié."}, status=400)

                    request.session['last_activity'] = {'name': 'ajouter_notes', 'time': timezone.now().isoformat()}
                    messages.success(request, "Les notes ont été importées avec succès.")
                    return JsonResponse({'status': 'success'})
                else:
                   messages.error(request, f"Le fichier comporte une erreur: {str(e)}")
                   return JsonResponse({'Erreur lors du traitement': str(e)}, status=500)
                    
            except Exception as e:
                messages.error(request, f"Erreur lors du traitement du fichier: {str(e)}")
                print(f"erreur 2: {str(e)}")
                return JsonResponse({'Erreur lors du traitement': str(e)}, status=500)
        except Exception as e:
            messages.error(request, f"Désolé, une erreur est survenue: {str(e)}")
            print("erreur 3")
            return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
        
def charger_notes(request):
    classes = Classe.objects.all()
    matieres = Matiere.objects.all()
    theme = request.session['theme'] if 'theme' in request.session else ''
    return render(request, 'charger_notes.html', {'classes': classes, 'matieres': matieres,'theme':theme})

def filtrer_matieres(request):
    """
        Cette fonction prend en parametre une classe et/ou un semestre contenu dans la requete puis, 
        retourne la liste des matieres correspondantes grace à une reponse json
    """
    classe_name = request.GET.get('classe')
    semestre = request.GET.get('semestre')
    if classe_name and semestre:
        classe = Classe.objects.get(name=classe_name)
        matieres = Matiere.objects.filter(classe=classe, semestre=semestre, active=True)
    elif classe_name:
        classe = Classe.objects.get(name=classe_name)
        matieres = Matiere.objects.filter(classe=classe, active = True)
    elif semestre:
        matieres = Matiere.objects.filter(semestre=semestre, active=True)
    else:
        matieres = Matiere.objects.none()

    data = [{"id":mat.id, "name":mat.name} for mat in matieres]

    return JsonResponse(data, safe=False) # si safe=True, Django s'attend a recevoir un dictionnaire exclusivement. Pour safe=False, on peut passer un objet autre qu'un dictionnaire (une liste, un tuple par exemple)

def ajouter_matiere(request):
    """Ajoute une matière dans la base de données"""
    theme = request.session['theme'] if 'theme' in request.session else ''
    if request.method == 'POST':
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        filiere_responsable = responsable.filiere
        form = MatiereForm(request.POST,filiere_responsable=filiere_responsable)
        if form.is_valid():
            form.save()
            request.session['last_activity'] = {'name': 'ajouter_matiere', 'time': timezone.now().isoformat()}
            messages.success(request, 'Matière ajoutée avec succès')
            return render(request, 'ajouter_matieres.html', {'form': MatiereForm(),'theme':theme})
    else:
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        filiere_responsable = responsable.filiere
        form = MatiereForm(filiere_responsable=filiere_responsable)
    return render(request, 'ajouter_matieres.html', {'form': form,'theme':theme})

def ajouter_matieres(request):
    """Ajoute des matieres reçues à partir d'un fichier excel"""
    if request.method == 'POST':
        data = request.POST
        fichier = request.FILES.get('fichier_excel')

        className = data.get('classe', "")
        try:
            classe = Classe.objects.get(name=className)
            try:
                df = pd.read_excel(fichier)
                for index, row in df.iterrows():
                    name = row['matiere']
                    semestre = row['semestre']
                    credit = row['credit']
                    Matiere.objects.create(name=name, semestre=semestre, classe=classe, credit=credit)
                request.session['last_activity'] = {'name': 'ajouter_matieres', 'time': timezone.now().isoformat()}
                messages.success(request, "Les matières ont été ajoutées avec succès.")
                return JsonResponse({'status': 'success'})
            except Exception as e:
                messages.error(request, f"Erreur lors du traitement du fichier: {str(e)}")
                return JsonResponse({'Erreur lors du traitement': str(e)}, status=500)
        except Exception as e:
            messages.error(request, f"Désolé, une erreur est survenue: {str(e)}")
            return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
        
@is_login
def charger_matieres(request):
    """Renvoie sur une page html où sera fourni un fichier excel contenant 
    une liste des matières à ajouter dans la base"""
    theme = request.session['theme'] if 'theme' in request.session else ''
    try: 
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        classes = responsable.get_classes_by_filiere()
    except Responsable.DoesNotExist:
        classes = Classe.objects.none()
        messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
    
    return render(request, 'charger_matieres.html', {'classes': classes,'theme':theme})

"""
   
    #matieres_etudiant = notes_etudiant.values_list('matiere__name', flat=True).distinct()

    # On recupere uniquement les noms des matieres pour la classe de l'etudiant et on range dans une liste
    #matieres_classe = Matiere.objects.filter(classe=etudiant.classe).values_list('name', flat=True)

    matieres_classe = Matiere.objects.filter(classe=etudiant.classe)

        # Initialiser un defaultdict pour les matières par semestre
    matieres_par_semestre = defaultdict(lambda: defaultdict(lambda: ['/','/']))

   
"""

@is_login
def notes_etudiants(request):
    """
        Cette fonction permet d'afficher les notes des etudiants
        selon la classe et le semestre selectionner par l'utilisateur
    """
    theme = request.session['theme'] if 'theme' in request.session else ''

    etudiant = Etudiant.objects.get(pk=request.session['user_pk'])
    notes_etudiant = Note.objects.filter(etudiant=etudiant)

    nb_messages_non_lus = Message.objects.filter(etudiant=etudiant, lu=False).count()

    # Obtenir les annees passees à l'école
    annees_scolaires = etudiant.years_at_school
    
    # Recuperer l'annee selectionner, depuis la methode GET effectuee
    selected_year = request.GET.get('annee_scolaire', etudiant.annee_scolaire_en_cours)
    
    # filtrer les notes de l'etudiant suivant le selected year
    notes_etudiant_annee = notes_etudiant.filter(annee_scolaire=selected_year)

    class_annee = notes_etudiant_annee.first().classe if notes_etudiant_annee.exists() else None
    
    # On recupere toutes les matieres pour la classe correspondante
    matieres_classe = Matiere.objects.filter(classe=class_annee)
    
    # On range les matieres suivant le semestre
    matieres_semestre1 = matieres_classe.filter(semestre='semestre1')
    matieres_semestre2 = matieres_classe.filter(semestre='semestre2')
    
    # On initialise les dictionnaires pour stocker les notes par matière et semestre
    notes_par_semestre1 = {}
    notes_par_semestre2 = {}

    # On remplit les notes pour les matieres du semestre 1
    for matiere in matieres_semestre1:
        note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note1').first()
        note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note2').first()
        moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
        notes_par_semestre1[matiere] = [
            note1.note if note1 else '/', 
            note2.note if note2 else '/',
            moyenne.moyenne if moyenne else '/'
        ]

    #  On remplit les notes pour les matieres du semestre 2
    for matiere in matieres_semestre2:
        note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note1').first()
        note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note2').first()
        moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
        notes_par_semestre2[matiere] = [
            note1.note if note1 else '/', 
            note2.note if note2 else '/',
            moyenne.moyenne if moyenne else '/'
        ]

    return render(request, 'note_etudiant.html', {
        'etudiant': etudiant,
        'notes_semestre1': notes_par_semestre1,
        'notes_semestre2': notes_par_semestre2,
        'annees_scolaires': annees_scolaires,
        'selected_year': selected_year,
        'classe_annee': class_annee,
        'theme': theme,
        'nb_message': nb_messages_non_lus
    })


@is_login
def ajouter_responsable(request):
    if request.method == 'POST':
        form = ResponsableForm(request.POST)
        theme = request.session['theme'] if 'theme' in request.session else ''
        if form.is_valid():
            form.save()
            messages.success(request, 'Responsable ajouté avec succès')
            return render(request, 'ajouter_responsable.html', {'form': ResponsableForm(),'theme':theme})
    else:
        form = ResponsableForm()
        theme = request.session['theme'] if 'theme' in request.session else ''
    return render(request, 'ajouter_responsable.html', {'form': form,'theme':theme})

@is_login
def liste_etudiants(request):

    if request.method == 'GET':

        # On filtre les classes qui correspondent au responsable connecté
        try: 
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            classes = responsable.get_classes_by_filiere()
        except Responsable.DoesNotExist:
            classes = Classe.objects.none()
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")

        # La liste des statuts possibles
        statuts=["En cours de formation","Diplômé","Exclus"]

        annee=request.GET.get('annee')
        statut=request.GET.get('statut')
        classe=request.GET.get('classe')

        selected_classe = classe
        selected_annee = annee
        selected_statut = statut

        all_notes = Note.objects.all()
        annees=[note.annee_scolaire for note in all_notes ]
        annees = list(set(annees))

        theme = request.session['theme'] if 'theme' in request.session else ''
        cycle1 = ['AS1','AS2','ISEP1','ISEP2','ISEP3','ISE1-eco','ISE1-maths','ISE2']
        responsable = Responsable.objects.get(pk=request.session['user_pk'])

        type = "" # Pour un affichage personnalisé niveau front-end

        if not (annee or statut or classe ):
            etudiants = responsable.get_etudiants_by_filiere()
 
        elif annee and statut and classe:
            if statut == "En cours de formation":
                note_filter=Note.objects.filter(annee_scolaire=annee,classe__name=classe)
                etudiant_filter=[note.etudiant for note in note_filter]
                
                # On retire les doublons
                etudiant_filter=list(set(etudiant_filter))
                etudiants_resp = responsable.get_etudiants_by_filiere()
                
                # Liste finale
                etudiants_p=[]
                for etudiant in etudiant_filter:
                    if etudiant.statut==statut:
                        etudiants_p.append(etudiant)

                for classe_responsable, eleves in etudiants_resp.items():
                    if classe_responsable==classe:
                        etudiants={classe: etudiants_p}
                        break

            elif statut == "Diplômé":
                etudiants_p = Etudiant.objects.filter(annee_diplomation=annee,classe__name=classe)
                etudiants_resp = responsable.get_etudiants_by_filiere()
                type="de diplomation"
                for classe_responsable, eleves in etudiants_resp.items():
                    if classe_responsable==classe:
                        etudiants={classe: etudiants_p}
                        break
            elif statut == "Exclus":
                etudiants_p = Etudiant.objects.filter(annee_exclusion=annee,classe__name=classe)
                etudiants_resp = responsable.get_etudiants_by_filiere()
                type = "d' exclusion"
                for classe_responsable, eleves in etudiants_resp.items():
                    if classe_responsable==classe:
                        etudiants={classe: etudiants_p}
                        break

        elif annee and classe:
            note_filter=Note.objects.filter(annee_scolaire=annee,classe__name=classe)
            etudiant_filter=[note.etudiant for note in note_filter]
            etudiants_p=list(set(etudiant_filter))
            etudiants_resp = responsable.get_etudiants_by_filiere()
            for classe_responsable, eleves in etudiants_resp.items():
                if classe_responsable==classe:
                    etudiants={classe: etudiants_p}
                    break

        elif annee and statut:
            if statut=="En cours de formation":

                note_filter=Note.objects.filter(annee_scolaire=annee)
                etudiant_filter=[note.etudiant for note in note_filter]
                etudiant_filter=list(set(etudiant_filter))
                etudiants_resp = responsable.get_etudiants_by_filiere()
                etudiants_p=[]
                for etudiant in etudiant_filter:
                    if etudiant.statut==statut:
                        etudiants_p.append(etudiant)

                for classe_responsable, eleves in etudiants_resp.items():
                    liste_etudiants=[]
                    for etudiant in etudiants_p:
                        if classe_responsable==etudiant.classe.name:
                            liste_etudiants.append(etudiant)
                    etudiants_resp[classe_responsable]=liste_etudiants

                etudiants=etudiants_resp
            elif statut == "Diplômé":
                etudiants_p = Etudiant.objects.filter(annee_diplomation=annee)
                etudiants_resp = responsable.get_etudiants_by_filiere()
                for classe_responsable, eleves in etudiants_resp.items():
                    liste_etudiants=[]
                    for etudiant in etudiants_p:
                        if classe_responsable==etudiant.classe.name:
                            liste_etudiants.append(etudiant)
                    etudiants_resp[classe_responsable]=liste_etudiants
                
                type = "de diplomation"

                etudiants=etudiants_resp
            elif statut == "Exclus":
                etudiants_p = Etudiant.objects.filter(annee_exclusion=annee,classe__name=classe)
                etudiants_resp = responsable.get_etudiants_by_filiere()
                for classe_responsable, eleves in etudiants_resp.items():
                    liste_etudiants=[]
                    for etudiant in etudiants_p:
                        if classe_responsable==etudiant.classe.name:
                            liste_etudiants.append(etudiant)
                    etudiants_resp[classe_responsable]=liste_etudiants

                etudiants=etudiants_resp

                type = "d' exclusion"

        elif statut and classe:
            etudiants_resp = responsable.get_etudiants_by_filiere()
            for classe_responsable, eleves in etudiants_resp.items():
                if classe_responsable == classe:
                    liste_etudiants=[]
                    for etudiant in eleves:
                        if etudiant.statut==statut:
                            liste_etudiants.append(etudiant)
                    etudiants_resp={classe_responsable:liste_etudiants}
                    break

            etudiants=etudiants_resp

            if statut == "Diplômé":
                type = "de diplomation"
            elif statut == "Exclus":
                type = "d' exclusion"
        

        elif annee:
            note_filter=Note.objects.filter(annee_scolaire=annee)
            etudiant_filter=[note.etudiant for note in note_filter]
            etudiant_filter=list(set(etudiant_filter))
            etudiants_resp = responsable.get_etudiants_by_filiere()
                                 

            for classe_responsable, eleves in etudiants_resp.items():
                liste_etudiants=[]
                for etudiant in etudiant_filter:
                    if classe_responsable==etudiant.classe.name:
                        liste_etudiants.append(etudiant)
                etudiants_resp[classe_responsable]=liste_etudiants

            etudiants=etudiants_resp

        elif statut:
            etudiants_resp = responsable.get_etudiants_by_filiere()
            for classe_responsable, eleves in etudiants_resp.items():
                liste_etudiants=[]
                for etudiant in eleves:
                    if etudiant.statut==statut:
                        liste_etudiants.append(etudiant)
                etudiants_resp[classe_responsable]=liste_etudiants

            etudiants=etudiants_resp
            
            if statut == "Diplômé":
                type = "de diplomation"
            elif statut == "Exclus":
                type = "d' exclusion"
            
    return render(request, 'liste_etudiants.html', {'etudiants': etudiants,'theme':theme,'cycle1':cycle1,'statuts':statuts,'classes':classes,'annees':annees,'selected_annee':selected_annee,'selected_classe':selected_classe,'selected_statut':selected_statut,'type':type})


@is_login
def Accueil_responsable(request):
    theme = request.session['theme'] if 'theme' in request.session else ''
    responsable = Responsable.objects.get(id = request.session['user_pk'])
    classes = responsable.get_classes_by_filiere()
    username = responsable.name or ""

    last_login = responsable.last_login
    last_activity = request.session.get('last_activity')
    return render(request, 'responsable.html',{'classes':classes,'last_login':last_login,'last_activity':last_activity,'theme':theme,'username':username})

@is_login
def Accueil_etudiant(request):
    theme = request.session['theme'] if 'theme' in request.session else ''
    etudiant = Etudiant.objects.get(pk = request.session['user_pk'])
    username = etudiant.name or ""
    nb_messages_non_lus = Message.objects.filter(etudiant=etudiant, lu=False).count()

    last_login = etudiant.last_login
    last_activity = request.session.get('last_activity')
    return render(request, 'etudiant.html',{'last_login':last_login,'last_activity':last_activity,'theme':theme,'username':username,'nb_message':nb_messages_non_lus})


@is_login
def ajouter_etudiant(request):
    """Ajoute un étudiant dans la base"""
    responsable = Responsable.objects.get(pk=request.session['user_pk'])
    classes = responsable.get_classes_by_filiere()
    username = responsable.name or ""
    if request.method == 'POST':
        form = EtudiantForm(request.POST)
        if form.is_valid():
            matricule = form.cleaned_data['matricule']
            if Etudiant.objects.filter(matricule=matricule).exists():
                messages.error(request, f"L'étudiant avec le matricule {matricule} existe déjà.")
                return render(request, 'ajouter_etudiant.html', {'form': form})
            else:
                form.save()
                request.session['last_activity'] = {'name':'ajouter_etudiant','time':timezone.now().isoformat()}
                messages.success(request, f"Etudiant {form.cleaned_data['name']} ajouté avec succès !!!")
                return render(request, 'ajouter_etudiant.html', {'form': EtudiantForm()})
        else:
            messages.error(request, form.errors)
            return render(request, 'ajouter_etudiant.html', {'form': form,'classes':classes})
    else:
        form = EtudiantForm()
        return render(request, 'ajouter_etudiant.html', {'form': form,'classes':classes,'username':username})

@is_login
def charger_etudiants(request):
    """Renvoie sur une page html où sera fourni un fichier excel contenant 
    une liste d'étudiants à ajouter dans la base"""
    try: 
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        classes = responsable.get_classes_by_filiere()
       
    except Responsable.DoesNotExist:
        classes = Classe.objects.none()
        messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
    username = responsable.name if responsable.name else ""
    return render(request, 'charger_etudiants.html', {'classes': classes,'username':username})


def ajouter_etudiants(request):
    """
        Cette fonction permet d'ajouter les étudiants dans la base de données
        à partir d'un fichier Excel que l'on va parcourir ligne par ligne
    """
    if request.method == 'POST':
        try:
            data = request.POST
            fichier = request.FILES.get('fichier_excel')

            classeName = data.get('classe', "")
            annee_inscription = data.get('annee_inscription', "")

            classe = Classe.objects.get(name=classeName)

            try:
                df = pd.read_excel(fichier)
                for index,row in df.iterrows():
                    matricule = row['matricule']

                    if Etudiant.objects.filter(matricule=matricule).exists():
                        messages.error(request, f"L'étudiant avec le matricule {matricule} existe déjà.")
                        return JsonResponse({'status': 'error', 'message': f"L'étudiant avec le matricule {matricule} existe déjà."}, status=404)
                    else:
                        Etudiant.objects.create(
                            matricule=matricule,
                            name=row['nom'],
                            classe=classe,
                            email=row['email'],
                            password=make_password(row['matricule']),
                            sexe = row['sexe'],
                            nationalite = row['nationalite'],
                            date_naissance = row['date_naissance'],
                            annee_inscription=annee_inscription
                        )
                request.session['last_activity'] = {'name':'ajouter_etudiants','time':timezone.now().isoformat()}
                print('erreur à ce niveau0')
                messages.success(request, "Les étudiants ont été ajoutés avec succès.")
                return JsonResponse({'status': 'success'})
            except Exception as e:
                print('erreur à ce niveau')
                messages.error(request, f"Erreur lors du traitement du fichier: {str(e)}")
                return JsonResponse({'Erreur lors du traitement': str(e)}, status=500)
        except Exception as e:
            print('erreur à ce niveau2')
            messages.error(request, f"Désolé, une erreur est survenue: {str(e)}")
            return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    
def modifier_matiere(request):
    """
        Cette fonction permet de modifier une matière spécifier lorsque la méthode est de type POST
        Pour les methodes GET, elle va filtrer la base et renvoyer les matières correspondantes, en fonction 
        de informations qui seront fournies par la méthode GET
    """
    if request.method == 'POST':
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        filiere_responsable = responsable.filiere
        form = MatiereForm(request.POST,filiere_responsable=filiere_responsable)
        matiere = Matiere.objects.get(name=form.data['name'],classe = form.data['classe'])
        form2 = MatiereForm(request.POST, instance=matiere)
        if form2.is_valid():
            form2.save()
            request.session['last_activity'] = {'name':'modifier_matiere','time':timezone.now().isoformat()}
            messages.success(request, "Matière modifiée avec succès.")
            # Rediriger pour éviter un rechargement du POST
            return redirect('modifier_matiere')  
        else:
            # Si le formulaire n'est pas valide, on renvoie la même page avec les erreurs
            try: 
                responsable = Responsable.objects.get(pk=request.session['user_pk'])
                # On filtre les classes pour le select du formulaire html
                classes = responsable.get_classes_by_filiere()
               
                
                semestres = [{'value':'semestre1','name':'Semestre 1'},
                            {'value':'semestre2','name':'Semestre 2'}]
            except Responsable.DoesNotExist:
                classes = Classe.objects.none()
                messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
                return render(request, 'modifier_matiere.html', {'classes': classes})
            else:   
                messages.error(request, "Erreur dans la modification de la matière.")
                return render(request,'modifier_matiere.html',{'form':form,'classes':classes,'semestres':semestres})
                
    elif request.method == 'GET':
        try: 
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            classes = responsable.get_classes_by_filiere()
            
            semestres = [{'value':'semestre1','name':'Semestre 1'},
                        {'value':'semestre2','name':'Semestre 2'}]
        except Responsable.DoesNotExist:
            classes = Classe.objects.none()
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
            return render(request, 'modifier_matiere.html', {'classes': classes})
        else:
            id_matiere = request.GET.get('id_matiere','')
            matiere = Matiere.objects.get(id=id_matiere) if id_matiere else None
            selected_matiere = (matiere.id if matiere else '')

            classe = request.GET.get('classe', '')
            classe = Classe.objects.get(name=classe) if classe else None
            selected_classe = (classe.name if classe else '')

            semestre = request.GET.get('semestre', '') 
            selected_semestre = semestre if semestre else ''

            if semestre and classe:
                matieres = Matiere.objects.filter(classe=classe, semestre=semestre, active=True)
            elif classe:
                matieres = Matiere.objects.filter(classe=classe, active=True)
            #elif semestre:
                #matieres = Matiere.objects.filter(semestre=semestre, active=True)
            else:
                matieres = Matiere.objects.none()
            
            # On recupere la filiere du responsable
            filiere_responsable = responsable.filiere

            # On charge les informations avant d'envoyer le formulaire
            form = MatiereForm(instance=matiere,filiere_responsable=filiere_responsable) if matiere else MatiereForm(filiere_responsable=filiere_responsable)

            return render(request, 'modifier_matiere.html', {'form':form,'matiere': matiere, 'classes': classes,'semestres':semestres, 'matieres': matieres,'selected_matiere':selected_matiere,'selected_classe':selected_classe,'selected_semestre':selected_semestre})
    
    # Au cas où la méthode n'est ni POST ni GET
    try:
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        filiere_responsable = responsable.filiere
        form = MatiereForm(filiere_responsable=filiere_responsable)
        return render(request, 'modifier_matiere.html',{'form':form})
    except Responsable.DoesNotExist:
        return HttpResponse('Méthode non autorisée')

def ajouter_classe(request):
    """Cette fonction permet d'ajouter une nouvelle classe dans la base de données"""
    if request.method == 'POST':
        form = ClasseForm(request.POST)
        if form.is_valid():
            classeName = form.cleaned_data['name']
            for classe in Classe.objects.all():
                if classe.name.lower() == classeName.lower():
                    messages.error(request, 'Cette classe existe déjà')
                    return render(request, 'ajouter_classe.html', {'form': form})
            form.save()
            request.session['last_activity'] = {'name':'ajouter_classe','time':timezone.now().isoformat()}
            messages.success(request, 'Classe ajoutée avec succès')
            return render(request, 'ajouter_classe.html', {'form': ClasseForm()})
        else:
            messages.error(request, form.errors)
            return render(request, 'ajouter_classe.html', {'form': form})
    else:
        form = ClasseForm()
    return render(request, 'ajouter_classe.html', {'form': form})

@is_login
def detail_etudiant(request, matricule):
    """Cette fonction permet d'afficher les informations d'un étudiant spécifique"""
    
    etudiant = Etudiant.objects.get(pk=matricule)
    notes_etudiant = Note.objects.filter(etudiant=etudiant)

    # Obtenir les annees passees à l'école
    annees_scolaires = etudiant.years_at_school
    
    # Recuperer l'annee selectionner, depuis la methode GET effectuee
    selected_year = request.GET.get('annee_scolaire', etudiant.annee_scolaire_en_cours)
    
    # filtrer les notes de l'etudiant suivant le selected year
    notes_etudiant_annee = notes_etudiant.filter(annee_scolaire=selected_year)

    class_annee = notes_etudiant_annee.first().classe if notes_etudiant_annee.exists() else None
    
    # On recupere toutes les matieres pour la classe correspondante
    matieres_classe = Matiere.objects.filter(classe=class_annee)
    
    # On range les matieres suivant le semestre
    matieres_semestre1 = matieres_classe.filter(semestre='semestre1')
    matieres_semestre2 = matieres_classe.filter(semestre='semestre2')
    
    # On initialise les dictionnaires pour stocker les notes par matière et semestre
    notes_par_semestre1 = {}
    notes_par_semestre2 = {}

    # On remplit les notes pour les matieres du semestre 1
    for matiere in matieres_semestre1:
        note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note1').first()
        note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note2').first()
        moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
        notes_par_semestre1[matiere] = [
            note1.note if note1 else '/', 
            note2.note if note2 else '/',
            moyenne.moyenne if moyenne else '/'
        ]

    #  On remplit les notes pour les matieres du semestre 2
    for matiere in matieres_semestre2:
        note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note1').first()
        note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note2').first()
        moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
        notes_par_semestre2[matiere] = [
            note1.note if note1 else '/', 
            note2.note if note2 else '/',
            moyenne.moyenne if moyenne else '/'
        ]

    return render(request, 'detail_etudiant.html', {
        'etudiant': etudiant,
        'notes_semestre1': notes_par_semestre1,
        'notes_semestre2': notes_par_semestre2,
        'annees_scolaires': annees_scolaires,
        'selected_year': selected_year,
        'classe_annee': class_annee
    })

@is_login
def modifier_etudiant(request,matricule):
    """
        Cette fonction permet de modifier un étudiant spécifié lorsque la méthode est de type POST
    """
    if request.method == 'POST':
        form = EtudiantUpdateForm(request.POST, instance=Etudiant.objects.get(pk=matricule))
        if form.is_valid():
            etudiant = Etudiant.objects.get(pk=matricule)
            password = form.data['password']
            if not check_password(password, etudiant.password):
                etudiant.password = make_password(password)
            form.save()
            request.session['last_activity'] = {'name':'modifier_etudiant','time':timezone.now().isoformat()}
            messages.success(request, f'Etudiant immatriculé {matricule} modifié avec succès')
            return redirect('detail_etudiant', matricule=matricule)
        else:
            messages.error(request, form.errors)
            return render(request, 'modifier_etudiant.html', {'form': form})
    else:
        form = EtudiantUpdateForm(instance=Etudiant.objects.get(pk=matricule))
        return render(request, 'modifier_etudiant.html', {'form': form})
    
@is_login
def modifier_etudiant2(request):
    theme = request.session.get('theme', 'light')
    if request.method == 'POST':
        form = EtudiantUpdatePasswordForm(request.POST, instance=Etudiant.objects.get(pk=request.session['user_pk']))
        if form.is_valid():
            etudiant = Etudiant.objects.get(pk=request.session['user_pk'])
            password = form.data['password']
            if not check_password(password, etudiant.password):
                etudiant.password = make_password(password)
            form.save()
            request.session['last_activity'] = {'name':'modifier_profil','time':timezone.now().isoformat()}
            messages.success(request, f'Vos informations ont été modifié avec succès')
            return redirect('notes_etudiants')
        else:
            messages.error(request, form.errors)
            return render(request, 'modifier_etudiant.html', {'form': form})
    else:
        form = EtudiantUpdatePasswordForm(instance=Etudiant.objects.get(pk=request.session['user_pk']))
        return render(request, 'etudiant_modifier_profil.html', {'form': form,'theme':theme})

@is_login  
def modifier_responsable(request):
    if request.method == 'POST':
        try:
            #responsable = Responsable.objects.get(pk=id_responsable)
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            form = ResponsableForm(request.POST, instance=responsable)
            if form.is_valid():
                password = form.data['password']
                if not check_password(password, responsable.password):
                    responsable.password = make_password(password)
                form.save()
                request.session['last_activity'] = {'name':'modifier_responsable','time':timezone.now().isoformat()}
                messages.success(request, f'Vos informations ont été modifié avec succès')
                return redirect('accueil_responsable')
            else:
                messages.error(request, form.errors)
                return render(request, 'modifier_responsable.html', {'form': form})
        except Responsable.DoesNotExist:
            messages.error(request, "Méthode non autorisée !!!")
            return redirect('accueil_responsable')
    else:
        try:
            #responsable = Responsable.objects.get(pk=id_responsable)
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            form = ResponsableForm(instance=responsable)
            return render(request, 'modifier_responsable.html', {'form': form})
        except Responsable.DoesNotExist:
            messages.error(request, "Méthode non autorisée !!!")
            return redirect('accueil_responsable')
        
def set_theme(request):
    """Permet de stocker le  thème appliqué dans la session"""
    if request.method == 'POST':
        data = json.loads(request.body)
        theme = data.get('theme', 'light')
        request.session['theme'] = theme
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'fail'}, status=400)

def send_message(request, matricule):
    responsable = Responsable.objects.get(pk=request.session['user_pk'])
    etudiant = Etudiant.objects.get(pk=matricule)
    if request.method == 'POST':
        form = OneMessageForm(request.POST)
        fichiers = request.FILES.getlist('fichier')
        if form.is_valid():
            message = form.save(commit=False)
            message.etudiant = Etudiant.objects.get(pk=matricule)
            message.responsable = responsable
            message.save()
            if fichiers:
                for fichier in fichiers:
                    FichiersJoints.objects.create(message=message, fichier=fichier)
            messages.success(request, 'Message envoyé avec succès')
            return render(request, 'send_message.html', {'form': OneMessageForm(),'etudiant':etudiant})
        else:
            messages.error(request, form.errors)
            return render(request, 'send_message.html', {'form': form, 'etudiant':etudiant})
    else:
        form = OneMessageForm()
        return render(request, 'send_message.html', {'form': form,'etudiant':etudiant})

@is_login
def delete_etudiant(request, matricule):
    try:
        etudiant = Etudiant.objects.get(pk=matricule)
        etudiant.delete()
        messages.success(request, f'Etudiant immatriculé {matricule} supprimé avec succès')
    except Etudiant.DoesNotExist:
        messages.error(request, "Aucun étudiant avec cet identifiant dans la base.")
    return redirect('liste_etudiants')

def mes_messages(request):
    theme = request.session['theme'] if 'theme' in request.session else ''
    etudiant = Etudiant.objects.get(pk=request.session['user_pk'])
    messages = Message.objects.filter(etudiant=etudiant,visible_etudiant=True).order_by('-created_at')
    nb_messages_non_lus = messages.filter(lu=False).count()

    return render(request, 'mes_messages.html', {'mes_messages': messages, 'nb_message':nb_messages_non_lus,'theme':theme})

def delete_message(request, message_id):
    """L'étudiant n'étant pas habileté à supprimer un message dans la base
    on va juste rendre le message qu'il souhaite supprimer invisible sur son interface"""
    try:
        message = Message.objects.get(pk=message_id)
        message.visible_etudiant = False
        message.save()
        messages.success(request, 'Message supprimé avec succès')
    except Message.DoesNotExist:
        messages.error(request, "Aucun message avec cet identifiant dans la base.")
    return redirect('mes_messages')

def ouvrir_fichier(request, fichier_joint_id):
    """
        Cette fonction permet d'ouvrir un fichier (si le navigateur le permet car pour certains types
        notamment les fichiers excel, certains navigateurs autorisent uniquement le telechargement)
    """
    try:
        fichier_joint = FichiersJoints.objects.get(pk=fichier_joint_id)
       
        file_path = fichier_joint.fichier.path

        # Définir le type de contenu en fonction de l'extension du fichier
        if file_path.endswith('.pdf'):
            content_type = 'application/pdf'
        elif file_path.endswith(('.jpg', '.jpeg', '.png', '.gif')):
            content_type = 'image/*'
        elif file_path.endswith('.txt'):
            content_type = 'text/plain'
        else:
            content_type = 'application/octet-stream'  # Type par défaut pour les fichiers binaires

        response = FileResponse(fichier_joint.fichier.open(), content_type=content_type)
        response['content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
    
        return response
           
    except FichiersJoints.DoesNotExist:
        messages.error(request, "Aucun fichier avec cet identifiant dans la base.")
        Http404("Aucun fichier avec cet identifiant dans la base.")

def enregistrer_fichier(request, fichier_joint_id):
    try:
        fichier_joint = FichiersJoints.objects.get(pk=fichier_joint_id)
        
        file_path = fichier_joint.fichier.path
        response = FileResponse(open(file_path, 'rb'))
        # quote permet de gerer les caractères spéciaux dans le nom du fichier
        response['content-Disposition'] = f'attachement; filename="{quote(os.path.basename(file_path))}"'
        return response
            
       
    except FichiersJoints.DoesNotExist:
        messages.error(request, "Aucun fichier avec cet identifiant dans la base.")
        Http404("Aucun fichier avec cet identifiant dans la base.")

def turn_to_lu(request, message_id):
    if request.method == 'GET':
        try:
            message = Message.objects.get(pk=message_id)
            message.lu = True
            message.save()
            nb_messages_non_lus = Message.objects.filter(etudiant=message.etudiant, lu=False).count()
            return JsonResponse({'status': 'success','nb_messages_non_lus':nb_messages_non_lus})
        except Message.DoesNotExist:
            return JsonResponse({'status': 'error','message':'Aucun message avec cet identifiant dans la base'}, status=404)
        
    else:
        return JsonResponse({'status':'error','message': 'Méthode non autorisée'}, status=405)
    
def messages_responsable(request):
    """Il s'agit de la view pour la page messages de 
    l'interface des responsables"""
    theme = request.session['theme'] if 'theme' in request.session else ''

    responsable = Responsable.objects.get(pk=request.session['user_pk'])
    messages = Message.objects.filter(responsable=responsable,visible_responsable=True).order_by('-created_at')
    return render(request, 'messages.html',{'messages_envoyes':messages,'theme':theme})

def responsable_cacher_message(request, message_id):
    """Cette fonction rend un message non disponible sur l'interface du responsable"""
    try:
        message = Message.objects.get(pk=message_id)
        message.visible_responsable = False
        message.save()
    except Message.DoesNotExist:
        messages.error(request, "Aucun message avec cet identifiant dans la base.")
    return redirect('messages')

def responsable_delete_message(request, message_id):
    """Cette fonction supprime le message spécifié dans la base"""
    try:
        message = Message.objects.get(pk=message_id)
        message.delete()
        messages.success(request, 'Message supprimé avec succès')
    except Message.DoesNotExist:
        messages.error(request, "Aucun message avec cet identifiant dans la base.")
    return redirect('messages')


def new_message(request):
    theme = request.session['theme'] if 'theme' in request.session else '' 
    if request.method == 'GET':
        try:
            classe_id = request.GET.get('classe_id')
            if classe_id:
                etudiants = Etudiant.objects.filter(classe__id=classe_id, statut='En cours de formation')
                selected_class = Classe.objects.get(pk=classe_id)
                
            else:
                etudiants = Etudiant.objects.none()
                selected_class = None
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            classes = responsable.get_classes_by_filiere()

            classes = Classe.objects.filter(name__in=classes)    
            return render(request, 'new_message.html', {'classes': classes,'etudiants':etudiants,'selected_class':selected_class,'theme':theme})
        
        except Responsable.DoesNotExist:
            classes = Classe.objects.none()
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
            return render(request, 'new_message.html', {'classes': classes})
    else:
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        sujet = request.POST.get('sujet')
        contenu = request.POST.get('contenu')
        fichiers = request.FILES.getlist('fichier') if 'fichier' in request.FILES else None
        classe_id = request.POST.get('classe_selected')
        classe = Classe.objects.get(pk=classe_id) if classe_id else None
        etudiants_selectionnes_id = request.POST.getlist('etudiants_selected') if 'etudiants_selected' in request.POST else None
        if etudiants_selectionnes_id:
            for etudiant_id in etudiants_selectionnes_id:
                etudiant = Etudiant.objects.get(pk=etudiant_id)
                message = Message.objects.create(responsable=responsable, etudiant=etudiant, sujet=sujet, message=contenu)
                if fichiers:
                    for fichier in fichiers:
                        FichiersJoints.objects.create(message=message, fichier=fichier)
            messages.success(request, 'Message envoyé avec succès au groupe d\'étudiants sélectionné')
        else:
            etudiants = Etudiant.objects.filter(classe=classe)
            for etudiant in etudiants:
                message = Message.objects.create(responsable=responsable, etudiant=etudiant, sujet=sujet, message=contenu)
                if fichiers:
                    for fichier in fichiers:
                        FichiersJoints.objects.create(message=message, fichier=fichier)
            messages.success(request, 'Message envoyé avec succès à toute la classe')

        return redirect('new_message')
        
def Upgrade_etudiant(request,matricule):
    """
        Cette fonction permet de mettre à jour un étudiant spécifique
    """
    try:
        etudiant = Etudiant.objects.get(pk=matricule) 
        etudiant.Upgrade_etudiant
        messages.success(request,f"étudiant envoyé en {etudiant.classe} avec le statut {etudiant.statut}") 
        return redirect('liste_etudiants')
    except Etudiant.DoesNotExist:
        messages.error(request, "Aucun étudiant avec cet identifiant dans la base.")
        return redirect('liste_etudiants')
    
def Upgrade_classe(request):
    """
        Cette fonction permet de mettre à jour les étudiants d'une classe
        Faire passer en classe superieur ou diplômer en fonction de la classe actuelle de l'étudiant
    """
    if request.method == 'POST':
        classe_name = request.POST.get('classe')
        exclude_students = request.POST.getlist('exclude_students')
        if classe_name:
            try:
                classe = Classe.objects.get(name=classe_name)
                etudiants = Etudiant.objects.filter(classe=classe, statut='En cours de formation')
                for etudiant in etudiants:
                    if etudiant.pk not in exclude_students:
                        etudiant.Upgrade_etudiant
                messages.success(request, f"étudiants de la classe {classe_name} ont été mis à jour avec succès")
                return redirect('liste_etudiants')
            except Classe.DoesNotExist:
                messages.error(request, "Aucune classe avec ce nom dans la base.")
                return redirect('liste_etudiants')
            
def Degrade_etudiant(request,matricule):
    """
        Cette fonction permet de mettre à jour un étudiant spécifique
    """
    try:
        etudiant = Etudiant.objects.get(pk=matricule) 
        etudiant.Degrade_etudiant
        messages.success(request,f"Etudiant {etudiant.name} reprend la classe {etudiant.classe.name}") 
        return redirect('liste_etudiants')
    except Etudiant.DoesNotExist:
        messages.error(request, "Aucun étudiant avec cet identifiant dans la base.")
        return redirect('liste_etudiants')
                        

@is_login
def Responsable_notes(request):
    """Cette fonction permet d'accéder à la page note de l'interface des responsables"""
    
    if request.method == 'GET':
        try: 
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            classes = responsable.get_classes_by_filiere()
        except Responsable.DoesNotExist:
            classes = Classe.objects.none()
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
        
        # On recupere le label de la classe envoyé depuis la méthode get (si disponible)
        class_name = request.GET.get('classe_name', None)
        # On recupere le matricule de l'etudiant envoyé depuis la methode get (si disponible)
        etudiant_id = request.GET.get('etudiant_id', None)

        theme = request.session['theme'] if 'theme' in request.session else ''
        

        if etudiant_id:
            # Si l'étudiant est présent, nos filtres s'interesseront principalement a ce dernier
            etudiant = Etudiant.objects.get(pk=etudiant_id)
            notes_etudiant = Note.objects.filter(etudiant=etudiant)

            selected_classe = etudiant.classe.name

            # On recupere la classe correspondante ainsi que les etudiants de cette en classe qui sont en formation
            classe = Classe.objects.get(name=class_name)
            etudiants = Etudiant.objects.filter(classe=classe, statut='En cours de formation')

            # Obtenir les annees passees à l'école
            annees_scolaires = etudiant.years_at_school
        
            # Recuperer l'annee selectionner, depuis la methode GET effectuee
            selected_year = request.GET.get('annee_scolaire', etudiant.annee_scolaire_en_cours)
        
            # filtrer les notes de l'etudiant suivant le selected year
            notes_etudiant_annee = notes_etudiant.filter(annee_scolaire=selected_year)

            class_annee = notes_etudiant_annee.first().classe if notes_etudiant_annee.exists() else None
            
            # On recupere toutes les matieres pour la classe correspondante
            matieres_classe = Matiere.objects.filter(classe=class_annee)
            
            # On range les matieres suivant le semestre
            matieres_semestre1 = matieres_classe.filter(semestre='semestre1')
            matieres_semestre2 = matieres_classe.filter(semestre='semestre2')
            
            # On initialise les dictionnaires pour stocker les notes par matière et semestre
            notes_par_semestre1 = {}
            notes_par_semestre2 = {}

            # On remplit les notes pour les matieres du semestre 1
            for matiere in matieres_semestre1:
                note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note1').first()
                note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre1', type_note='note2').first()
                moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
                notes_par_semestre1[matiere] = [
                    note1.note if note1 else '/', 
                    note2.note if note2 else '/',
                    moyenne.moyenne if moyenne else '/'
                ]

            #  On remplit les notes pour les matieres du semestre 2
            for matiere in matieres_semestre2:
                note1 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note1').first()
                note2 = notes_etudiant_annee.filter(matiere=matiere, semestre='semestre2', type_note='note2').first()
                moyenne = Moyenne.objects.filter(etudiant=etudiant, matiere=matiere, annee_scolaire=selected_year).first()
                notes_par_semestre2[matiere] = [
                    note1.note if note1 else '/', 
                    note2.note if note2 else '/',
                    moyenne.moyenne if moyenne else '/'
                ]

            return render(request, 'notes.html', {
                'selected_etudiant': etudiant,
                'notes_semestre1': notes_par_semestre1,
                'notes_semestre2': notes_par_semestre2,
                'annees_scolaires': annees_scolaires,
                'selected_year': selected_year,
                'classe_annee': class_annee,
                'classes': classes,
                'selected_classe': selected_classe,
                'etudiants': etudiants,
                'theme': theme
            })
        
        elif class_name:
            # On recupere la classe correspondante ainsi que les etudiants de la classe
            classe = Classe.objects.get(name=class_name)
            etudiants = Etudiant.objects.filter(classe=classe, statut='En cours de formation')
            return render(request, 'notes.html', {'etudiants': etudiants, 'classes': classes,'selected_classe':classe.name,'theme':theme})
        else:
            # On retourne uniquement les classes correspondantes au responsable si ni une classe, ni un étudiant n'est selectionné
            return render(request, 'notes.html', {'classes': classes,'theme':theme})



def modifier_et_telecharger_excel(request):

    if request.method == 'GET':
        # Récupérer les données de la requête 
        annee_scolaire = request.GET.get('annee_scolaire') if request.GET.get('annee_scolaire') else None
        classe_name = request.GET.get('classe') if request.GET.get('classe') else None

        # Vérifier si les données sont présentes
        if annee_scolaire and classe_name:
            
            try:
                classe = Classe.objects.get(name=classe_name)
            except Classe.DoesNotExist:
                classe = None

            
            def collecter_donnees_semestre(semester):
                """Cette fonction va permettre de collecter les notes relatives a un semestre données
                pour la classe et l'année considérée"""
        
                # Récupérer les notes de l'étudiant pour l'année scolaire et la classe spécifiées
                notes = Note.objects.filter( annee_scolaire=annee_scolaire, classe=classe, semestre =semester).order_by('matiere_id')
                
                notes1= [note for note in notes if note.type_note == 'note1']
            
                notes2 = [note for note in notes if note.type_note == 'note2']
            
                # On recupere les matieres correpondantes
                matieres1 = [note.matiere for note in notes]
                
                # On supprime les doublons
                matieres = []
                for matiere in matieres1:
                    if matiere not in matieres:
                        matieres.append(matiere)
                
                
                etudiants1=[note.etudiant for note in notes]
            
                etudiants=[]

                for etudiant in etudiants1:
                    if etudiant not in etudiants:
                        etudiants.append(etudiant)


                # Récupérer les moyennes des etudiants pour l'année scolaire et la classe spécifiées
                moyennes1 = [matiere.moyennes.all() for matiere in matieres]
            
                moyennes = []
                for listemoyenne in moyennes1:
                    for moyenne in listemoyenne:
                        if moyenne not in moyennes:
                            moyennes.append(moyenne)

            
                # C'est dans ce dictionnaire que seront stockées toutes les informations a remplir dans le fichier excel
                grand_dict = dict()

                for etudiant in etudiants:
                    grand_dict[etudiant] = dict()
                    for matiere in matieres:
                        grand_dict[etudiant][matiere] = {
                            'note1': None,
                            'note2': None,
                            'moyenne': None,
                        }
                        for note in notes1:
                            if note.etudiant == etudiant and note.matiere == matiere:
                                grand_dict[etudiant][matiere]['note1'] = note.note
                        for note in notes2:
                            if note.etudiant == etudiant and note.matiere == matiere:
                                grand_dict[etudiant][matiere]['note2'] = note.note
                        moyenne = matiere.moyennes.filter(etudiant=etudiant).first()
                        if moyenne:
                            grand_dict[etudiant][matiere]['moyenne'] = moyenne.moyenne
                    
                return {'grand_dict':grand_dict,'matieres':matieres}



            # Chemin vers le fichier Excel existant
            chemin_fichier = 'modeles/modele_ise3_eval.xlsx'
            
            # Ouvrir le fichier Excel
            #wb = openpyxl.load_workbook(chemin_fichier)

            app = xw.App(visible=False) # Pour que en ouvrant le fichier excel, ce ne soit pas visible
            wb  = app.books.open(chemin_fichier)

            def remplir_feuille_semestre(semester,grand_dict,matieres):
                """Cette fonction permet de remplir le fichier excel a partir des informations collectées au prealable"""

                #ws = wb[semester]  # Sélectionner la feuille active (ou spécifiez une feuille spécifique avec `wb['Nom de la feuille']`)
                ws = wb.sheets[semester] # avec xlwings
                
                ligne = 5

                for etudiant, data in grand_dict.items():
                    """
                    ws.cell(row=ligne, column=1, value=etudiant.pk)
                    ws.cell(row=ligne, column=2, value=etudiant.name)
                    ws.cell(row=ligne, column=3, value=etudiant.sexe)
                    ws.cell(row=ligne, column=4, value=etudiant.date_naissance)
                    ws.cell(row=ligne, column=5, value=etudiant.nationalite)
                    ws.cell(row=ligne, column=6, value=etudiant.annee_inscription)
                    ws.cell(row=ligne, column=7, value=etudiant.email)
                    ws.cell(row=ligne, column=8, value=etudiant.heure_absence)
                    """
                    ws.range(f'A{ligne}').value = etudiant.pk
                    ws.range(f'B{ligne}').value = etudiant.name
                    ws.range(f'C{ligne}').value = etudiant.sexe
                    ws.range(f'D{ligne}').value = etudiant.date_naissance
                    ws.range(f'E{ligne}').value = etudiant.nationalite
                    ws.range(f'F{ligne}').value = etudiant.annee_inscription
                    ws.range(f'G{ligne}').value = etudiant.email
                    ws.range(f'H{ligne}').value = etudiant.heure_absence
                    
                    # Remplir les notes pour chaque matière
                    col_base = 9  # Début des colonnes pour les matières
                    for matiere in matieres:
                        """
                        ws.cell(row=ligne, column=col_base, value=data[matiere]['note1'])
                        ws.cell(row=ligne, column=col_base + 1, value=data[matiere]['note2'])
                        ws.cell(row=ligne, column=col_base + 2, value=data[matiere]['moyenne'])
                        """
                        ws.range(f'{xw.utils.col_name(col_base)}{ligne}').value = data[matiere]['note1']
                        ws.range(f'{xw.utils.col_name(col_base + 1)}{ligne}').value = data[matiere]['note2']
                        ws.range(f'{xw.utils.col_name(col_base + 2)}{ligne}').value = data[matiere]['moyenne']
                        col_base += 3  # Passer aux colonnes suivantes pour la prochaine matière

                    ligne += 1

            # On recupere les informations du semestre1
            grand_dict1 = collecter_donnees_semestre('semestre1')['grand_dict']
            matieres1 = collecter_donnees_semestre('semestre1')['matieres']

            # On recupere les informations du semestre 2
            grand_dict2 = collecter_donnees_semestre('semestre2')['grand_dict']
            matieres2 = collecter_donnees_semestre('semestre2')['matieres']
            
            # On remplit le fichier pour les smestres 1 et 2
            remplir_feuille_semestre('semestre1',grand_dict1,matieres1)
            remplir_feuille_semestre('semestre2',grand_dict2,matieres2)

            # Enregistrer les modifications dans un fichier temporaire
            #fichier_temp = 'modeles/fichier_temp.xlsx'
            #wb.save(fichier_temp)

            #fichier_temp = io.BytesIO()
            #wb.save(fichier_temp)
            #fichier_temp.seek(0)

            # Créer une réponse HTTP pour le téléchargement
            nom_fichier = f"recapitulatif_{annee_scolaire}_{classe.name}_.xlsx"

            # Préparez la réponse HTTP
            """response = HttpResponse(
                fichier_temp,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={nom_fichier}'

            """
            # Sauvegarder dans un fichier temporaire
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                fichier_temp = tmp.name
            wb.save(fichier_temp)  # Sauvegarde
            wb.close()  # Ferme le classeur
            xw.apps.active.quit()  # Quitte Excel
            
            with open(fichier_temp, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={nom_fichier}'

            # Supprimer le fichier temporaire après le téléchargement
            os.remove(fichier_temp)

            return response
        else:
            # Gérer le cas où les données sont manquantes
            messages.error(request, 'Veuillez sélectionner une année scolaire et une classe.')
            return redirect('liste_etudiants')  # Rediriger vers la page des notes





def interaction_page(request):
    results = []
    if request.method == 'POST':
        data = request.POST.get('data_field')
        # Ajouter ou filtrer des données dans la base de données
        #YourModel.objects.create(name=data, value=data)
        return redirect('interaction_page')  # Rediriger vers la même page pour rafraîchir les résultats

    #results = YourModel.objects.all()
    return render(request, 'modele.html', {'results': results})


@is_login
def dashboard(request):

    try:
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        classes = responsable.get_classes_by_filiere()
    except Responsable.DoesNotExist:
        classes = Classe.objects.all()
        messages.error(request, 'Accès non autorisé !!!')
    
    enseignants = Enseignants.objects.count()

    moyennes_etudiants = []
    for etudiant in Etudiant.objects.all():
        moyenne_etudiant = Note.objects.filter(etudiant=etudiant,annee_scolaire=etudiant.annee_scolaire_en_cours).aggregate(Avg('note'))['note__avg'] or 0
        moyennes_etudiants.append({
            'name':etudiant.name,
            'classe':etudiant.classe.name,
            'moyenne':moyenne_etudiant
        })
    
    

    # Filtrer par ordre décroissant de la moyenne
    moyennes_etudiants.sort(key=lambda x: x['moyenne'], reverse=True)

    def construction_histogramme_moyenne(moyennes_etudiants):
        """
        # Création de l'histogramme
        plt.figure(figsize=(8, 6))
        plt.hist([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants], bins=5, edgecolor='black')
        plt.title('Histogramme des moyennes')
        plt.xlabel('Moyenne')
        plt.ylabel('Nombre d\' étudiants ')

        # Sauvegarde dans un buffer mémoire
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()  # Fermer la figure après génération

        histogram_base64 = base64.b64encode(image_png).decode('utf-8')
        """

        # Création de l'histogramme interactif
        moyennes = [moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]
        trace = go.Histogram(x=moyennes, nbinsx=8, marker=dict(color='rgb(0,123,255)'))
        layout = go.Layout(title='Histogramme des Moyennes',
                        xaxis=dict(title='Moyenne'),
                        yaxis=dict(title='Nombre d\'étudiants'))

        fig = go.Figure(data=[trace], layout=layout)

        # Génère le HTML du graphique
        graph_div = opy.plot(fig, auto_open=False, output_type='div')

        return graph_div
    
    # compter le nombre d'étudiants avec une moyenne inférieure à 12
    def Compter_moyennes_inf_12(moyennes_etudiants):
        moyennes_inf_12 = [moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants if moyenne_etudiant['moyenne'] < 12]
        return len(moyennes_inf_12)



    if request.method == "GET":
        classe_name = request.GET.get('classe')
        semestre = request.GET.get('semestre')

        if classe_name:
            classe = Classe.objects.get(name=classe_name)
            etudiants_classe = Etudiant.objects.filter(classe=classe).count()
            etudiants_filles = Etudiant.objects.filter(classe=classe, sexe="F").count()
            etudiants_garcons = Etudiant.objects.filter(classe=classe, sexe="M").count()
            moyenne = Note.objects.filter(classe=classe).aggregate(Avg('note'))['note__avg'] or 0

            moyennes_etudiants = [] # vider la liste
            for etudiant in Etudiant.objects.filter(classe=classe):
                moyenne_etudiant = Note.objects.filter(etudiant=etudiant,annee_scolaire=etudiant.annee_scolaire_en_cours).aggregate(Avg('note'))['note__avg'] or 0
                moyennes_etudiants.append({
                    'name':etudiant.name,
                    'classe':etudiant.classe.name,
                    'moyenne':moyenne_etudiant
                })


            if semestre:
                moyennes_etudiants = [] # vider la liste
                moyenne = Note.objects.filter(classe=classe, semestre=semestre).aggregate(Avg('note'))['note__avg'] or 0
                for etudiant in Etudiant.objects.filter(classe=classe):
                    moyenne_etudiant = Note.objects.filter(etudiant=etudiant,semestre=semestre,annee_scolaire=etudiant.annee_scolaire_en_cours).aggregate(Avg('note'))['note__avg'] or 0
                    moyennes_etudiants.append({
                        'name':etudiant.name,
                        'classe':etudiant.classe.name,
                        'moyenne':moyenne_etudiant
                    })
            

            # Filtrer par ordre décroissant de la moyenne
            moyennes_etudiants.sort(key=lambda x: x['moyenne'], reverse=True)

           
            # Statistiques moyennes
            statistiques_moyennes = {
                "moyenne": np.mean([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]),
                "min": np.min([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
                "max": np.max([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
                "median": np.median([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
                "sd": np.std([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
                "histogram": construction_histogramme_moyenne(moyennes_etudiants),
                
            }

            return render(request, 'dashboard.html', {'classes': classes, 
                                                      'selected_classe': classe_name, 
                                                      'etudiants_classe': etudiants_classe,
                                                      'eleves_count': etudiants_classe,
                                                      'girls_count': etudiants_filles,
                                                      'boys_count': etudiants_garcons,
                                                      'moyenne_generale':moyenne,
                                                      'selected_semestre':semestre,
                                                      'enseignants':enseignants,
                                                      'meilleures_performances':moyennes_etudiants[:5],
                                                      'stats_moyennes': statistiques_moyennes,
                                                      'moyennes_inf_12': Compter_moyennes_inf_12(moyennes_etudiants),
                                                      })
    
    # statistiques moyennes concernant toute l'école
    statistiques_moyennes = {
        "moyenne": np.mean([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]),
        "min": np.min([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
        "max": np.max([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
        "median": np.median([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
        "sd": np.std([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]) if len(moyennes_etudiants)>0 else None,
        "histogram": construction_histogramme_moyenne(moyennes_etudiants),
    }
    
    context = {
        'eleves_count': Etudiant.objects.count(),
        'girls_count': Etudiant.objects.filter(sexe="F").count(),
        'boys_count': Etudiant.objects.filter(sexe="M").count(),
        'moyenne_generale': np.mean([moyenne_etudiant['moyenne'] for moyenne_etudiant in moyennes_etudiants]),
        'nouvelles_notes': Note.objects.filter(created_at__gte=timezone.now()-timedelta(days=7)).count(),
        'recent_notes': Note.objects.order_by('-created_at')[:5],
        'classes':classes,
        'enseignants':enseignants,
        'meilleures_performances':moyennes_etudiants[:5],
        'stats_moyennes': statistiques_moyennes,
        'moyennes_inf_12': Compter_moyennes_inf_12(moyennes_etudiants),
    }
    return render(request, 'dashboard.html', context)

@is_login
def RespoTimeTable(request):
    if request.method == "GET":
        days = ["lundi","mardi","mercredi","jeudi","vendredi","samedi"]
        try:
            responsable = Responsable.objects.get(pk=request.session['user_pk'])
            classes = responsable.get_classes_by_filiere()
  
            selected_classe = request.GET.get('classe')
            selected_periode = request.GET.get('periode')

            # informations sur les emplois de temps
            def get_emploi_data(classe_name, periode=None):
                try:
                    classe = Classe.objects.get(name=classe_name)
                    emplois = EmploiDuTemps.objects.filter(classe=classe).order_by('-created_at')
                    periodes = [emploi.periode for emploi in emplois]
                    if periode:
                        emploi = EmploiDuTemps.objects.filter(classe=classe, periode=periode).order_by('-created_at').first()
                        return emploi, periodes, emplois
                    return [], periodes, emplois
                except (Classe.DoesNotExist, EmploiDuTemps.DoesNotExist):
                    return None, None, None

            # classe et periode fournies
            if selected_classe and selected_periode:
                emploi, periodes, emplois = get_emploi_data(selected_classe, selected_periode)
                if emploi:
                    programmations = emploi.programmation_cours.all()
                    
                    print(programmations)
                    return render(request, 'time_table.html', {
                        "classes": classes,
                        "periodes": periodes,
                        "selected_classe": selected_classe,
                        "emploi_de_temps": emploi,
                        "selected_periode": selected_periode,
                        "programmations": programmations,
                        "numero_emploi": len(emplois),
                        "days": days
                    })
                else:
                    messages.error(request, "Aucun emploi du temps trouvé pour cette classe et cette période.")
                    return render(request, 'time_table.html', {
                        "classes": classes,
                        "selected_classe":selected_classe,
                        "selected_periode":selected_periode,
                        "days":days
                        })
            
            # classe fournie sans periode
            elif selected_classe:
                _, periodes, emplois = get_emploi_data(selected_classe)
                if emplois:
                    return render(request, 'time_table.html', {
                        "classes": classes,
                        "periodes": periodes,
                        "selected_classe": selected_classe,
                        "numero_emploi": len(emplois),
                        "days":days
                    })
                else:
                    messages.error(request, "Aucune classe trouvée pour ce label.")
                    return render(request, 'time_table.html', {
                        "classes": classes,
                        "selected_classe":selected_classe,
                        "days":days
                        })
        
        except Responsable.DoesNotExist:
            classes = Classe.objects.none()
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
            return render(request, 'time_table.html', {"classes": classes})

    # si la requete n'est pas GET
    return render(request, 'time_table.html', {"classes": classes,"days":days})

   
@is_login
def MakeTimeTable(request):
    days = ["lundi","mardi","mercredi","jeudi","vendredi","samedi"]
    
    try: 
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        classes = responsable.get_classes_by_filiere()
    except Responsable.DoesNotExist:
        classes = Classe.objects.none()
        messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")

    if request.method == "GET":
        selected_classe = request.GET.get('classe') if request.GET.get('classe') else None
        selected_semester = request.GET.get('semestre') if request.GET.get('semestre') else None
        if selected_classe and selected_semester:
            try:
                nb_emploi =  EmploiDuTemps.objects.filter(classe__name=selected_classe, semestre=selected_semester).count() if EmploiDuTemps.objects.filter(classe__name=selected_classe, semestre=selected_semester).exists() else 0
                numero_emploi = nb_emploi + 1
                classe = Classe.objects.get(name=selected_classe)
                matieres = Matiere.objects.filter(classe=classe, semestre=selected_semester)
                return render(request, 'make_time_table.html',{"days":days,"numero_emploi":numero_emploi,"classes":classes,"matieres":matieres,"selected_classe":selected_classe,"selected_semestre":selected_semester})
            except Matiere.DoesNotExist:
                matieres = Matiere.objects.none()
                messages.error(request, "Aucune matière trouvée pour cette classe et ce semestre.")
                return render(request, 'make_time_table.html',{"days":days,"classes":classes})
        else:
            matieres = Matiere.objects.none()
            messages.error(request, "Veuillez sélectionner une classe et un semestre.")
            return render(request, 'make_time_table.html',{"days":days,"classes":classes})
    else:

        return render(request, 'make_time_table.html',{"days":days,"classes":classes})
    
@is_login
def SaveTimeTable(request):
    if request.method == "POST":
        data = request.POST # pour les données envoyées par le formulaire
        classe_name = data.get('classe',"")
        semestre = data.get('semestre', "")
        periode = data.get('periode', "").strip()
        emploi_json = data.get('emploi_de_temps', "[]") # recupère la chaîne json

        try:
            emploi_de_temps = json.loads(emploi_json) # Transforme en liste python
        except json.JSONDecodeError:
            emploi_de_temps = []

        if emploi_de_temps:
            print(emploi_de_temps[0])
            try:
                classe = Classe.objects.get(name=classe_name)
                time_table = EmploiDuTemps.objects.create(classe=classe, semestre=semestre, periode=periode) # On crée un nouvel emploi de temps
                time_table.save()

                for programmation in emploi_de_temps:
                    jour = programmation.get("jour", "").strip()
                    numero_ligne = programmation.get("numero_ligne", "").strip()
                    matiere_name = programmation.get("matiere", "").strip()
                    enseignant = programmation.get("enseignant", "").strip()
                    horaire = programmation.get("horaire", "").strip() 
                  
                    if numero_ligne:
                        try:
                            numero_jour = int(numero_ligne.split("_")[1])
                        except (IndexError, ValueError):
                            numero_jour = None
                    else:
                        numero_jour = None
                    
                        
                    if matiere_name:
                        matiere = Matiere.objects.get(name=matiere_name)

                        # Création du programme
                        programme = Programmation_cours.objects.create(matiere=matiere, horaire=horaire, jour=jour, numero = numero_jour, emploi_du_temps = time_table)
                        programme.save()
                        print(f"programme enregistré avec matiere {matiere}")
                                    
                    else:
                        continue
                    
                print("Emploi de temps sauvegardé")
                return JsonResponse({"success": "Emploi de temps sauvegardé"})

            except Exception as e:
                print(e)
                return JsonResponse({"error": "Erreur lors de l'enregistrement de l'emploi de temps"})

        else:
            print("Aucun emploi de temps trouvé")
        return JsonResponse({"error": "Aucun emploi de temps envoyé"})
    
@is_login
def DeleteTimeTable(request, classe, periode):
    if request.method == "POST":
       classe_name = classe
       periode_name = periode
       try:
           classe = Classe.objects.get(name=classe_name)
           emploi = EmploiDuTemps.objects.get(classe=classe, periode=periode_name)
           emploi.delete()
           messages.success(request, "Emploi de temps supprimé avec succès")
           return redirect('emploi_de_temps')
       except EmploiDuTemps.DoesNotExist:
            messages.error(request, "Emploi de temps non trouvé")
            return redirect('emploi_de_temps')

def DeleteTimeTableForModify(request):
    classe_name = request.GET.get('classe')
    periode_name = request.GET.get('periode')
    try:
        classe = Classe.objects.get(name=classe_name)
        emploi = EmploiDuTemps.objects.get(classe=classe, periode=periode_name)
        emploi.delete()
       
        return JsonResponse(data=["ok"], safe=False)
    except EmploiDuTemps.DoesNotExist:
        messages.error(request, "Emploi de temps non trouvé")
        return JsonResponse(data=["error"], safe=False)
       
@is_login
def EditTimeTable(request, classe, periode):
    days = ["lundi","mardi","mercredi","jeudi","vendredi","samedi"]

    # Vérification de l'authentification
    try:
        responsable = Responsable.objects.get(pk=request.session['user_pk'])
        classes = responsable.get_classes_by_filiere()
    except Responsable.DoesNotExist:
        classes = Classe.objects.none()
        messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")

    # Traitement de la méthode GET
    if request.method == "GET":
        selected_classe = classe
        selected_periode = periode
        try:
            classe = Classe.objects.get(name=selected_classe)
            emploi = EmploiDuTemps.objects.get(classe=classe, periode=selected_periode)
            programmations = emploi.programmation_cours.all()
            programmations_dict = {}
            for prog in programmations:
                print(prog)
                programmations_dict[f"{prog.jour}_{prog.numero}"] = {
                    "matiere": str(prog.matiere.name),
                    "enseignant": str(prog.matiere.enseignant),
                    "horaire": prog.horaire
                }
            semestre = emploi.semestre
            matieres = Matiere.objects.filter(classe=classe, semestre=semestre)
            return render(request, 'edit_time_table.html', {"days":days, "classes":classes, "matieres":matieres, "selected_classe":selected_classe, "selected_periode":selected_periode, "programmations":programmations, "emploi":emploi,"programmations_dict":programmations_dict,"semestre":semestre})
        except EmploiDuTemps.DoesNotExist:
            messages.error(request, "Emploi de temps non trouvé")
            return redirect('emploi_de_temps')


@is_login
def EtudiantTimeTable(request):
    if request.method == "GET":
        days = ["lundi","mardi","mercredi","jeudi","vendredi","samedi"]
        try:
            etudiant = Etudiant.objects.get(pk=request.session['user_pk'])
            nb_messages_non_lus = Message.objects.filter(etudiant=etudiant, lu=False).count()
  
            selected_classe = etudiant.classe
            selected_periode = request.GET.get('periode')

            # informations sur les emplois de temps
            def get_emploi_data(classe_name, periode=None):
                try:
                    classe = Classe.objects.get(name=classe_name)
                    emplois = EmploiDuTemps.objects.filter(classe=classe).order_by('-created_at')
                    periodes = [emploi.periode for emploi in emplois]
                    if periode:
                        emploi = EmploiDuTemps.objects.filter(classe=classe, periode=periode).order_by('-created_at').first()
                        return emploi, periodes, emplois
                    return [], periodes, emplois
                except (Classe.DoesNotExist, EmploiDuTemps.DoesNotExist):
                    return None, None, None

            # classe et periode fournies
            if selected_classe and selected_periode:
                emploi, periodes, emplois = get_emploi_data(selected_classe, selected_periode)
                if emploi:
                    programmations = emploi.programmation_cours.all()
                    
                    print(programmations)
                    return render(request, 'etudiant_time_table.html', {
                        "periodes": periodes,
                        "selected_classe": selected_classe,
                        "emploi_de_temps": emploi,
                        "selected_periode": selected_periode,
                        "programmations": programmations,
                        "numero_emploi": len(emplois),
                        "days": days,
                        "nb_message":nb_messages_non_lus
                    })
                else:
                    messages.error(request, "Aucun emploi du temps trouvé pour cette classe et cette période.")
                    return render(request, 'etudiant_time_table.html', {
                        "selected_classe":selected_classe,
                        "selected_periode":selected_periode,
                        "days":days,
                        "nb_message":nb_messages_non_lus
                        })
            
            # classe fournie sans periode
            elif selected_classe:
                _, periodes, emplois = get_emploi_data(selected_classe)
                if emplois:
                    return render(request, 'etudiant_time_table.html', {
                        "periodes": periodes,
                        "selected_classe": selected_classe,
                        "numero_emploi": len(emplois),
                        "days":days,
                        "nb_message":nb_messages_non_lus
                    })
                else:
                    messages.error(request, "Aucune classe trouvée pour ce label.")
                    return render(request, 'etudiant_time_table.html', {
                        "selected_classe":selected_classe,
                        "days":days,
                        "nb_message":nb_messages_non_lus
                        })
        
        except Etudiant.DoesNotExist:
            messages.error(request, "Vous n'êtes pas autorisé à accéder à cette page.")
            return redirect("connexion")

    # si la requete n'est pas GET
    return render(request, 'etudiant_time_table.html', {"classes": classes,"days":days,"nb_message":nb_messages_non_lus})


@is_login
def ajouter_enseignant(request):
    if request.method == 'POST':
        form = EnseignantForm(request.POST)
        theme = request.session['theme'] if 'theme' in request.session else ''
        if form.is_valid():
            form.save()
            messages.success(request, 'Enseignant ajouté avec succès')
            return render(request, 'ajouter_enseignant.html', {'form': EnseignantForm(),'theme':theme})
    else:
        form = EnseignantForm()
        theme = request.session['theme'] if 'theme' in request.session else ''
    return render(request, 'ajouter_enseignant.html', {'form': form,'theme':theme})
            

    